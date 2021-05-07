#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
__title__  = openpyxl操作Excel工具类
"""

import openpyxl
 

class ExcelUtil:
    workBook = None
    workSheet = None

    def load_excel(self, path):
        """
        加载Excel
        :param path: 需要打开的Excel的路径
        """
        self.workBook = openpyxl.load_workbook(path)
 
    def get_sheet_by_name(self, name):
        """
         获取sheet对象
        :param name: sheet名
        """
        self.workSheet = self.workBook.get_sheet_by_name(name)

    def get_sheet_by_index(self, index=0):
        """
        获取sheet对象
        :param index: sheet的索引
        """
        # 获取workBook里所有的sheet名 -> list
        sheet_names = self.workBook.get_sheet_names()
        # 根据索引获取指定sheet
        self.workSheet = self.workBook[sheet_names[index]]

    def get_cell_value(self, col, row):
        """
         获取cell的值
         :param col: 所在列
         :param row: 所在行
        """
        try:
            return self.workSheet.cell(column=col, row=row).value
        except BaseException as e:
            return None

    def get_cell_value_by_xy(self, str):
        """
        获取cell的值
        :param str: 坐标
        """
        try:
            return self.workSheet[str].value
        except BaseException as e:
            return None

    def get_sheet_rows(self):
        """
        获取最大行数
        """
        return self.workSheet.max_row

    def get_sheet_cols(self):
        """
        获取最大列数
        """
        return self.workSheet.max_column

    def write_data(self, row, col, value, path):
        """
        写入数据
        """
        try:
            self.workSheet = self.workBook.active
            self.workSheet.cell(column=col, row=row, value=value)
            self.workBook.save(path)
        except BaseException as e:
            print(e)
            return None
 
    def get_excel_data(self):
        """
        获取表所有数据
        :return: list
        """
        # 方式一
        data_list = tuple(self.workSheet.values)
        # 方式二
        # data_list = []
        # for i in range(self.get_sheet_rows()):
         #     data_list.append(self.get_row_value(i + 2))
        return data_list

    def get_row_value(self, row):
        """
        获取某一行的内容
        :param row: 第几行 -> str  **从1开始**
        :return: list
        """
         # 方式一
        row_list = self.get_excel_data()[row]
        # 方式二
        # row_list = []
        # for i in self.workSheet[str(row + 1)]:
        #     row_list.append(i.value)
        return row_list

    def get_col_value(self, col='A'):
        """
        获取某一列的内容
        :param col: 第几列 -> str
        :return: list
        """
        col_list = []
        for i in self.workSheet[col]:
            col_list.append(i.value)
        return col_list

    def get_row_num(self, case_id):
        """
        获取行号
        :param case_id: 用例编号
        :return:
        """
        num = 1
        col_data = self.get_col_value()
        for data in col_data:
            if case_id == data:
                return num
            num += 1
        return 0


excelUtil = ExcelUtil()

if __name__ == '__main__':
    path = 'C:/Users/Sogou-SunShijiang/Desktop/考勤.xlsx'
    # 读取excel文件
    excelUtil.load_excel(path)
    # 获取某个sheet
    excelUtil.get_sheet_by_name("Sheet")
    excelUtil.get_sheet_by_index()
    # 获取某个cell的值
    data = excelUtil.get_cell_value(col=1, row=1)
    print(data)
    data = excelUtil.get_cell_value_by_xy("A3")
    print(data)
    # 获取sheet行数
    data = excelUtil.get_sheet_rows()
    print(data)
    # 获取sheet列数
    data = excelUtil.get_sheet_cols()
    print(data)
    # 获取某一行数据
    data = excelUtil.get_row_value(0)
    print(data)
    # 获取某一列数据
    data = excelUtil.get_col_value()
    print(data)
    # 写入数据
    excelUtil.write_data(row=9, col=1, value="test", path=path)
    # 获取全部数据
    data = excelUtil.get_excel_data()
    print(data)
    # 获取行号
    data = excelUtil.get_row_num('imooc_001')
    print(data)