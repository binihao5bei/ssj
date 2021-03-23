from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
#一、创建新表
wb=Workbook()
#二、打开已有的表
#wb=load_workbook('商品销售.xlsx')
#二、正在激活的工作表
ws1=wb.active
#三、插入行标题（表头）
wshead=['店铺ID','商品ID','标题','总销量','日期']
ws1.append(wshead)

#四、储存数据
#1.单元格插入数据
ws1['A2']='杨国福麻辣烫'
#2.插入一行数据
ws1.append(['张亮麻辣烫','102','什锦套餐',100,'2021-03-23'])
#3.插入带日期的数据
ws1['E2']=datetime.datetime.now().strftime("%Y-%m-%d")

#五、创建新的sheet表页
ws3=wb.create_sheet('sheet3')     #默认插到最后
ws2=wb.create_sheet('sheet2',1)   #插到指定位置，0为最开始，1为第二个

#六、选择sheet表页
#1.根据sheet页名称（方法一）
ws2=wb['sheet2']                   #常用
#2.根据sheet页名称(方法二）                 
ws3=wb.get_sheet_by_name('sheet3') #不常用，且会出现warning警告

#七、查看表名
#1.直接显示所有sheet表页名
print(wb.sheetnames)        
#2.遍历所有sheet表页
for i in wb:
	print(i.title)  

#八、修改sheet表页名
ws4=wb['sheet2']            #第一步：定位到要修改名称的sheet表页
ws4.title='sheet2_rename'   #第二步：直接给该表页的title赋新的值即可
wb.save('商品销售.xlsx')     #第三步：保存excel表（为了保存操作记录）

#九、访问单元格
#1.单一单元格访问
#（1）方法一：
c=ws1['B2']     #访问sheet表页：ws1的B2单元格
ws1['C2']='女神套餐'
c.value='101'   #给单元格的value赋值
#（2）方法二：
d=ws1.cell(column=4,row=2,value=90)
#（3）方法三：
for i in range(1,101):
	for j in range(1,101):
		ws4.cell(column=j,row=i,value='xx')		
#2.多单元格访问
#（1）切片访问多个单元格
cell_range=ws3['A1':'C2']   #访问sheet3表页的A1:C2确定的区域范围
#（2）通过行（列）访问
colC=ws3['C']               #访问sheet3表页的第C列
col_range=ws3['C:E']        #访问sheet3表页的C列:E列确定的区域范围
row10=ws3[10]               #访问sheet3表页的第10行
row_range=ws3[10:15]        #访问sheet3表页的第10行:第15行确定区域范围
#（3）通过指定的范围(行-->行）
for row in ws3.iter_rows(min_row=1,max_row=3,max_col=2):
	for cell in row:
		print(cell)
#（4）通过指定的范围(列-->列）
for col in ws3.iter_cols(min_col=1,max_col=3,max_row=2):
	for cell in col:
		print(cell)		
#（5）遍历所有
ws3['C9']='Hello World'
print(tuple(ws3.rows))      #遍历sheet3表页的所有行
print(tuple(ws3.columns))   #遍历sheet3表页的所有列

#十、保存数据
wb.save('商品销售.xlsx')

#十一、其他
#1.改变sheet标签页按钮背景颜色
ws3.sheet_properties.tabColor='1072BA'
#2.获取最大行和最大列
print(f'sheet3的最大行数为:{ws3.max_row}')
print(f'sheet3的最大列数为:{ws3.max_column}')
#3.获取每一行和每一列
for row in ws1.rows: #ws1.rows为生成器，里面是每一行数据，每一行又有一个tuple包裹
	for cell_row in row:
		print(cell_row.value) #按行将每行数据输出出来

for col in ws1.columns: #ws1.columns为生成器，里面是每一列数据，每一列又有一个tuple包裹
	for cell_col in col:
		print(cell_col.value) #按列将每列数据输出出来
#4.根据数字得到字母，根据字母得到数字
from openpyxl.utils import get_column_letter,column_index_from_string
#(1)根据列的数字得到字母
print(get_column_letter(2))   #输出第二列的字母：B（第二列即B列）
#(2)根据列的字母得到数字
print(column_index_from_string('B')) #输出B列的数字：2（B列即第二列）
#5.删除工作表
#(1)方法一：
#wb.remove(ws3)
#(2)方法二：
#del wb['sheet3']
#6.矩阵转换（行列倒置）
rows=[
	['Number','data1','data2'],
	[2,40,30],
	[3,40,25],
	[4,50,30],
	[6,25,5],
	[7,50,10]
]
for i in list(zip(*rows)):
	ws3.append(i)

#十二、设置单元格样式
#1.需要导入类
from openpyxl.styles import Font,colors,Alignment,PatternFill,Border,Side
#2.设置字体
#直接使用的cell的font属性
font_bold=Font(name='等线',size=24,italic=True,color='00FF00',bold=True)
fill_bold=PatternFill(fill_type='darkDown',fgColor='AACF91',bgColor='1874CD')
ws3['A1'].font=font_bold
ws3['A2'].fill=fill_bold
#3.设置对齐方式
#直接使用的cell的属性alignment  cell.alignment=Alignment(horizontal='',vertical='')
align=Alignment(horizontal='center',vertical='center')
for row in ws1.rows:
	for cell in row:
		cell.alignment=align
#4.设置行高和列宽
#设置ws1sheet页第一行的行高为40
#ws1.row_dimensions[1].height=40
#设置ws1sheet页第2列的列宽为30
#ws1.column_dimensions[2].width=30
#5.合并和拆分单元格
#(1)合并单元格
ws1.merge_cells('E1:E3')
ws1.merge_cells('F1:H3')
#(2)拆分单元格
ws1.unmerge_cells('F1:H3')
#6.设置边框
for row in ws1.rows:
	for cell in row:
		cell.border = Border(top = Side(border_style='thin',color='FF000000'), right = Side(border_style='thin', color='FF000000'), bottom = Side(border_style='thin', color='FF000000'),left = Side(border_style='thin', color='FF000000'))



wb.save('商品销售.xlsx')


'''----------------------案例↓--------------------------'''
'''
from openpyxl import Workbook
import datetime
from time import time
from random import choice
from openpyxl.utils import get_column_letter

wb=Workbook()
ws=wb.active
ws.append(['Date','title','A_Z'])
for i in range(500):
	dati=datetime.datetime.now().strftime("%Y-%m-%d")
	title=str(time())
	a_z=get_column_letter(choice(range(1,50)))
	ws.append([dati,title,a_z])
wb.save('test.xlsx')
'''
'''----------------------案例↑--------------------------'''