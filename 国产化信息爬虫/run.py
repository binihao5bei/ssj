 # coding=utf-8
import datetime
import time
import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles  import Font, colors, Alignment,Border,Font,Side,PatternFill
import os

# ===web网站站点定义
from web_ukylin import Ukylin
from web_kylin import Kylin
from web_deepin import Deepin
from web_xinchuang import Xinch
from web_shurufa import Shurufa
from web_gcczxt import Gcczxt
#from web_IThome import IThome
from web_uos import UOS
# ===微信爬虫站点定义
from wx_search import WX_search

# ===微博爬虫站点定义
from wb_ukylin import Wb_ukylin
from wb_neokylin import Wb_neokylin
from wb_uos import Wb_uos

# 自定义函数
from func import htmlmake, sendmailnow, add_db, add_wiki, tablemake

# ===站点类清单
weblist = []
# 网页爬虫
#weblist += [Ukylin]
#weblist += [Ukylin,Kylin,Deepin,WX_search,Wb_ukylin,Wb_uos,Xinch,Shurufa,Gcczxt]
#weblist += [Ukylin,Deepin,WX_search,Wb_ukylin,Wb_uos,Xinch,Shurufa,Gcczxt]
weblist += [Deepin,WX_search,Wb_ukylin,Wb_uos,Xinch,Shurufa,Gcczxt]
#weblist += [Ukylin,Kylin,Deepin,UOS,Wb_ukylin,Wb_neokylin,Xinch,Shurufa,Gcczxt]
# 微信爬虫
#weblist += [Wx_ukylin]

# 初始化参数
F_datalists = []
nowdate = datetime.datetime.now().strftime('%Y-%m-%d')
# 每日最大发布记录条数
maxnum = 30

# ===测试用自定义日期
#nowdate = '2021-01-01'

# 遍历网站类列表
for webclass in weblist:
    # 类定义Kylin
    web = webclass(nowdate)
    # 数据获取
    F_datalists += web.get_data()

#列表根据里面的字典key（title）去重
datalists=[]
values=[]
for d in F_datalists:
    if d['title'].split('】')[1] not in values:
        datalists.append(d)
    values.append(d['title'].split('】')[1])

# 打印最终结果
#print(datalists)
print("==总共抓取到：" + str(len(datalists)) + "条记录。")


# 发邮件
if __name__ == '__main__':
    if len(datalists) > 0:
        mail_msg = htmlmake(maxnum, datalists)
        nowdate = datetime.datetime.now().strftime('%Y-%m-%d')
        
        #1.发送邮件
        sendmailnow(["ImeToB@sogou-inc.com"], "【信创今日最新消息】" + str(nowdate), mail_msg)
        #sendmailnow(["sunshijiang@sogou-inc.com"], "【信创今日最新消息】" + str(nowdate), mail_msg)
        print('1.已成功发送到ToB群邮件！')


        #2.发布到wiki
        pageConnect = tablemake(datalists)
        add_wiki(str(nowdate), pageConnect)
        print('2.已成功发布到wiki中！')
        
        
        

        #3.储存到数据库
        time=str(nowdate).replace('-','_')
        conn=pymysql.connect(host='10.129.157.42',port=3306,user='root',password='123456',database='sogou_vifereo')
        cursor=conn.cursor()
        
        sql1='''create table if not exists {} (
            TITLE varchar(255) not null,
            URL varchar(255),
            TIME date);'''.format(time)

        cursor.execute(sql1)
        conn.commit()


        sql2='''insert into {} (TITLE,URL,TIME) values(%s,%s,%s);'''.format(time)
        for i in datalists:
            cursor.execute(sql2,(i['title'],i['wurl'],i['wdate'])) 

        conn.commit()
        conn.close()
        print('3.已成功储存到数据库中！')
        
        
        
        
        #4.储存到excel表中
        wb=Workbook()     
        #wb=load_workbook('test.xlsx')
        ws=wb.active

        wshead=['标题','链接']
        ws.append(wshead)
        ws.row_dimensions[1].height=15
        ws.column_dimensions['A'].width=80
        ws.column_dimensions['B'].width=100
        font=Font('宋体',bold=True)
        fille = PatternFill('solid', fgColor="FFBB00")
        
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].font=font
        ws['B1'].font=font
        ws['A1'].fill=fille
        ws['B1'].fill=fille

        for i in datalists:
            ws.append([i['title'],i['wurl']])
            
        for row in ws.rows:
            for cell in row:
                cell.border = Border(top = Side(border_style='thin',color='FF000000'), right = Side(border_style='thin', color='FF000000'), bottom = Side(border_style='thin', color='FF000000'),left = Side(border_style='thin', color='FF000000'))

        wb.save(f'C:/Users/Sogou-SunShijiang/Desktop/FTP/国产化信息爬取源代码/data/{nowdate}.xlsx')
       
        if os.path.exists(f'C:/Users/Sogou-SunShijiang/Desktop/FTP/国产化信息爬取源代码/data/{nowdate}.xlsx'):
            print('4.excel表创建成功')
        else:
            print('4.未找打该Excel表')
        

