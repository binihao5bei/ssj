import requests
import selenium
from selenium import webdriver
import time
import datetime
import json
from interval import Interval
from urllib.parse import urlencode
from openpyxl import Workbook
#from openpyxl.styles import Font,Color,Alignment,Border,side
from openpyxl.styles import Font,colors,Alignment,PatternFill,Border,Side
import tkinter as tk
from tkinter import ttk
from tkinter import scrolledtext
from tkinter import Menu


# Create instance
win = tk.Tk()
sw=win.winfo_screenwidth()
sh=win.winfo_screenheight()
ww=220
wh=200
x=(sw-ww)/2
y=(sh-wh)/2
win.geometry('%dx%d+%d+%d'%(ww,wh,x,y))
win.title("考勤信息一键导出")  

tabControl = ttk.Notebook(win)         
tab1 = ttk.Frame(tabControl)           
tabControl.add(tab1, text='考勤')      
tabControl.pack(expand=1, fill="both") 

mighty = ttk.LabelFrame(tab1, text=' 考勤导出 ')
mighty.grid(column=0, row=0, padx=8, pady=4)

def click_me(): 
    #action.configure(text=number1_chosen.get()+"-"+number2_chosen.get())
    month_ip=number2_chosen.get()+"-"+number3_chosen.get()
    user=number1_chosen.get()
    '''--------------------------------------------------------------------------------------------------'''
    headers = {
        'Host' : 'oa.sogou-inc.com',
        'Referer' : 'https://oa.sogou-inc.com/statics/attd/web2/index.html',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36 Edg/89.0.774.54'
    }

    #根据日期返回星期几
    def get_week_day(date):
      week_day_dict = {
        0 : '星期一',
        1 : '星期二',
        2 : '星期三',
        3 : '星期四',
        4 : '星期五',
        5 : '星期六',
        6 : '星期日',
      }
      day = date.weekday()
      return week_day_dict[day]

    cookies1={"AUTH_COOKIE_KEY58": "dK99Wt571v6xkSzU5Za8JKKbyxGDZQBbPTwkUcbkZfiqMvEmLc9crfbUGT4ymBeKtxbr5jbdsU7v5MrkdH2WyYHZ9h3t5aW6txtytgFec611P8sEA76cfBzxjMarTSFEGDSqkJWBcnH2wnjcTegEEvihJG34JJYF8nxVGZeWWKi2MHZKzW5tb7zkHPS4h4kmcYcYRoibaNu4Ywmvu41HVquP6vgLfMZcxDzU",
              "JSESSIONID": "aaav3fXn_Sw8QdRfvtyHx",
              "AUTH_COOKIE_KEY": "eyJsb2dpbklkIjoic3Vuc2hpamlhbmciLCJuYW1lIjoi5a2Z5LiW5rGfIiwicmVxdWVzdElwIjoiMTAuMTI5LjE1OC45MyIsInNpZ24iOiJmNDFmODU4NTAwZTdiMzJmNmZjYzcxY2Y0NDY2NmQ2YWVlNzVjYzE3YWVjODg0MWE2NzAxMmZmYTc5YzNkYjI2IiwidGltZSI6MTYxNjM4OTE.",
              "SESSIONID": "bed43ab1-55e4-467e-bcfa-4bbe6057cb0a",
              "PAID": "spAPysJh5jMMznevRAJut8VyVWydQy27Dnq1H81Rf1q9QYqmC3CQQbazPdTemNm1crMdTM3RiqRDYpw2WDK2yC563qpWc8wCyeSjPStGjBoN6EpyuEiEN6vzMWdWCACwvojK1dfgVGhz3Rys7capYP8oTdeoJGtua"
              }
    cookies2={"UM_distinctid":"177d1c97fef10-0a9a8f52afccdf-73e356b-2a3000-177d1c97ff08aa",
              "_ga":"GA1.2.683675609.1614132380",
              "PAID":"3cHDDpPKsQVwyJByRR18ZSdmpy4EHScPQfMd5HCSn9As4kqKmDbKzCcdSnzrSapyVXf2SXmSKkQTnopxS3vB1r8QLzhUPkgriex24TxMYcMcGMTUD27T4m5gujW16NN98sJxAKUNJC9DbNfos8Gka4BNo7jjksn",
              "AUTH_COOKIE_KEY":"eyJsb2dpbklkIjoibWF5dXhpbmciLCJuYW1lIjoi6ams5a6H5pifIiwicmVxdWVzdElwIjoiMTAuMTI5LjE1Ni4xNDkiLCJzaWduIjoiYzdlMmNiYWJjNzE5MDJlNDI1NTY3NDI0YmNkNzk2YTNmMTFkOTM3NTM5NGRiNzJlMDQ0MTk4YzZjNWQ5YWUwNiIsInRpbWUiOjE2MTQ2NTQ2",
              "AUTH_COOKIE_KEY58":"2s7Lxui9mdZA2rPDf96RDdvxVPtSHkgvYoWtpKUNkw3x1cPCd7mid4UTxY4TNRviBAdqqsEicqYovA8eEkSC7ZoVrekZuUdZFYD3b6Rn1bTcQ3jj6hBV4mzJd16vNR3m9adiTJuQviKEtmbQRBKLSCd9jxjY2DWGUSaYupiX9bCHiyb8ZLEFoXy3ya1EZyoRW6MCe2Y3j1dGsd1pwKWb6b7kQJo5encQSp",
              "JSESSIONID":"aaaFdeP5LPrbw_dyT6WFx",
              "SESSIONID":"36ec7a20-0a35-4563-b17e-6558d6050784"
              }
    cookies3={"_ga":"GA1.2.661641063.1598410615",
              "UM_distinctid":"1780fcb66b9520-094c0d97b429d4-7a1437-1fa400-1780fcb66ba71",
              "PAID":"spAPysJh5jMMyuEnyUq4AV1XdqufVJ7tXWVj2wzG77xY1fi3s2RTNa5575d9PhpobGwbzETzwkfvrygZKdnLzfs8Y8bc2dsvK3BnkkcRrNtdm8xaVs8KBoH6PcqHbudVNXSWy7nZausG6mNVp8uVMg8HVtjzcVKSU",
              "AUTH_COOKIE_KEY":"eyJsb2dpbklkIjoibHVvaGFpeWFuZyIsIm5hbWUiOiLnvZfmtbfmtIsiLCJyZXF1ZXN0SXAiOiIxMC4xMjkuMTU3LjExMiIsInNpZ24iOiJiOWVjOGY3ZjA0NmVhYmUyMDYxMmQ0YWU1N2NmZTRjZjJjMmZkODJjMTg5MjNmZDJjNjRkMzA1MTYwNzk5YTI4IiwidGltZSI6MTYxNjA3MDk.",
              "AUTH_COOKIE_KEY58":"dK99Wt571v6xkSzU58neb6W2tUAQUkcoDdW7pKACHqHi5AQgYHtyobo1MFJjBW4HtfV7KDu4GtUSS54wXFjvNqzVy9x4vrXPzmCKLuWMwywCD9YHL25vpcHRJjyjFo8o8w2vr6joXUjTV1E6qLWtySrr5zQj9sXdFZnLSsGorp6fgxjMTxidAnh1Y8iyDJoT4Ty4f48R1bqnLQyunXDYkGKuXB3wbFDyxYEQ",
              "JSESSIONID":"aaaUsl7qGWv_JbC_xvfHx",
              "IPLOC":"ZZ",
              "SESSIONID":"51c852fa-7cab-4af0-8df4-49bd0a745d0b"
        }
    cookies4={"_ga":"GA1.2.853454607.1574680949",
              "UM_distinctid":"17858a4d33c10a-06951b1f517e588-5c140258-100200-17858a4d33d15b",
              "_gid":"GA1.2.1666183524.1616394902",
              "PAID":"CWiYwa9YHUimdpsBKSZiGgsgGnmAotiBWmbbw9MN1AJxAhc2QbTzWxQaw1Ahm1nBmJfvh9uZyq8KPoB2xSy9GiCwhXy7G3dRdU55LQLUzM6j9mRaXCsQ7Cu38HaME9dq76pn54TJDUiLbbMjeiFH2Tt29PnzcQ84",
              "AUTH_COOKIE_KEY":"eyJsb2dpbklkIjoieWV4aW5neGluZyIsIm5hbWUiOiLlj7bmmJ_mmJ8iLCJyZXF1ZXN0SXAiOiIxMC4xMjkuMjAuMTcxIiwic2lnbiI6ImEwYWM0OWYxZmFmNDFkZDQ3ZTU1MmFmYjk0NWI0Zjk0MzQxNDhhODQ1ZGFiM2E5OTA5ZGVhNTRhODE2MGVkOWEiLCJ0aW1lIjoxNjE2Mzk1Mg..",
              "AUTH_COOKIE_KEY58":"9EB17medhzLvsAy4sWMwqQp2YKMVdQcay6kQ47KvNMbpuF3617PuY74rPbmDoEV7mR1kN2w85Z7wZbJW1adSUyHiQbUGmBvEBxKMR545GGHjPDHV7SamQiqaLe9vAQH3PRuARqExXDZWR7tpSrdPq3wZK82XRZQUa47kTNrx2j9cdX8xR4Rk7frNc4nbfyZcWFcijtHQmaMrJ2pivyUJukMNverhgjSy1F2" ,
              "JSESSIONID":"aaaEA548qYg5Ddr_PQyHx",
              "SESSIONID":"aa037c28-2bd5-4667-8ed8-0e167ec0895b",
              "Asset.COOKIE":"PNxJGSK2tPHjudyjajPMMaTQgpMa9noNAsvBt2gsdBWEvnxekRjUiYDTubfznuiKFQMzZKVa3fmTzKADd2omJakdiwacWAJTCT2pzcAkZKx2b7Rya6PdkJEVNyGNx6FcZUVMso1zz4ZVqwzU8FRQi4pSdPgK3NG2r9qrdoC7MAQpaxSTaSgCqzzeXSXavgz8fjEYT37kZ17rwMe3v"
        }
    cookies5={"_ga":"GA1.2.853454607.1574680949", 
                "UM_distinctid":"17858a4d33c10a-06951b1f517e588-5c140258-100200-17858a4d33d15b", 
                "_gid":"GA1.2.1666183524.1616394902", 
                "PAID":"4sHAUS135usvsiFeuzPkSUFER1iiweyyNMSnvnRrgahfG5oENBvVNznNjvD4awUK7EaJFJwt2feKH8i3zjwQxGRtD7qERbEtLuwEaLcdL4GeP1Rs42mqFn2kEprcE8GdWhv9jyX4yFPnUwBGXeE9GUQF2AXrTKurxaY", 
                "AUTH_COOKIE_KEY":"eyJsb2dpbklkIjoid2FuZ3hpbnhpbjAxIiwibmFtZSI6IueOi-aso-asoyIsInJlcXVlc3RJcCI6IjEwLjEyOS4yMC4xNzEiLCJzaWduIjoiNTkyOThhZWMzMmQzN2M1N2E2Y2JkMWVkN2VkYjNiYTM2YmYzNjc5ODYzMjcxZDEwYmI5MmExNDUyODY3MzJlYiIsInRpbWUiOjE2MTYzOTg4", 
                "AUTH_COOKIE_KEY58":"3mH4xatR7Z3PK4mieYC3xzRdMsossJkvCQbv2eNh5R9vwRf7CeaCC3dG6KdJJobCdFNp4A8JBHvYpis7oYGnsrVrxJVTvUmw25GZFyeERxmFnzBUma4oo6zJRjwcsaro9gxVN2gMzMgDhzxn9517pJ4NfdYRJ9fzG9s8FNaU8qZ7Z1h2CDNHEoyH3yj2DxGTsAKF2eGefGRiHfyuRYAWk6ZZmFE8Dv5ccGbto6", 
                "JSESSIONID":"aaavggSrEgZjW_phu4yHx", 
                "SESSIONID":"06a8d7f6-1c84-492c-90aa-81a89f691306"
        }
    cookies6={"_ga":"GA1.2.853454607.1574680949",
            "UM_distinctid":"17858a4d33c10a-06951b1f517e588-5c140258-100200-17858a4d33d15b",
            "_gid":"GA1.2.1666183524.1616394902",
            "PAID":"3cHDDpPKsQVwyHqJEhAgqhnQPQMKeiARQv9Z87z2hUE4JgdKVne7kkSMmBKJMpSctJfaSMFQyjxFio9QL5vxGXgaQsjzviMXxgFPnCCkLT7KzWKaM2XqBU3Tk8UzNgYY3FYdcqApDffiZW979Zhf5cyPu4BfoCQ",
            "AUTH_COOKIE_KEY":"eyJsb2dpbklkIjoibG9uZ3NodWFpIiwibmFtZSI6Ium-meW4hSIsInJlcXVlc3RJcCI6IjEwLjEyOS4yMC4xNzEiLCJzaWduIjoiYWMxZTNjMzBlMDdiNjVmOWYyZjIzOWFkN2EyYzU3ZGZkY2Y4ZDk2OTMzNDI5NWYyODZlMzAxYjZjMTZjMWU3MyIsInRpbWUiOjE2MTYzOTc0MDg.",
            "AUTH_COOKIE_KEY58":"2FvSrYhYTLo7wEpxHFSjWVk855mMPAMtDdFAJcLBGoQoMjLrHdkeKXzHojcx6XxkCooGDY353JHP7aW5XZXeiBfM5LAq1gcUqke7drfuPfkwEcXV353MdKF5wwzY8VDV6QoPt6GEZd33dyqsFQwchhSE4ARcvBpqz7VVLSh2TBwK2ZZ6gABzJBMvRqVKAWa8w8GvJtZjFHskE17LcojfmjV6H85MBS",
            "JSESSIONID":"aaa4nnxh5mTHVa53XYyHx",
            "SESSIONID":"06ef8631-037a-486d-b16d-f1f42257d0da"}
    cookies7={"_ga":"GA1.2.853454607.1574680949",
              " UM_distinctid":"17858a4d33c10a-06951b1f517e588-5c140258-100200-17858a4d33d15b",
                " _gid":"GA1.2.1666183524.1616394902",
              "PAID":"CWiYwa9YHUimdAZgvEiNfwYjRexD6kJMR7MN4waFKyz4Q3q5hS2iArkjyTJwYGLkEAHRJaLZfoYmBucFzJNVDSA3331329gqEnyavAMak1ZJ1fWNEhFC5C4Ru8Nn8dmVg8n1vhk1swmd5BFDWmC17s27sVF5mW7n",
              "AUTH_COOKIE_KEY":"eyJsb2dpbklkIjoiYmFpeXVuamlhbiIsIm5hbWUiOiLnmb3kupHmtqciLCJyZXF1ZXN0SXAiOiIxMC4xMjkuMjAuMTcxIiwic2lnbiI6ImFmYjgwZDY3M2UxNjhhYWNkNGVkZDFiYmU1ZDI3NDE2YzVmY2NjNzdhNWU5NjVhZjQwZDZiZmUwYTZiZTM5Y2EiLCJ0aW1lIjoxNjE2Mzk4MQ..", "AUTH_COOKIE_KEY58":"9EB17medhzLvsAy4sBu7u771EkjudANBWkrdtqnGyEmjNHeuWBn9yBaRwv5RUbq9dYKhMeWUqGKLv8sh55nRET5ycnDgS3v9LDdFPrEQp4MqKc4irA32tywWUCL4J5Pd6ZPbKkX9BMH7aNgABYN33DWwcbvPXD32b3pXP4XBapKwUTKBgpgHWZ5N7gznzF5jwo6RXb2eUB5UAPkbEf5oqTxftZazoD2u2Ma",
             "SESSIONID":"717df9b6-059c-4edc-bbd2-6b38524f7ec2", 
            "JSESSIONID":"aaaIDt3Je_2wo9CMB1yHx"}
    cookies8={"_ga":"GA1.2.853454607.1574680949",
            "UM_distinctid":"17858a4d33c10a-06951b1f517e588-5c140258-100200-17858a4d33d15b",
            "_gid":"GA1.2.1666183524.1616394902",
            "PAID":"spAPysJh5jMN1Y9m6Q6rrZHvcx4G6JffwvkyvytsxLHEQoCJf7e8NWjVfQ73erTJGNAZeRDYTPPLB3wur4UK5sqUnmojinc3Jb3Ti3rs5bCdiEQrysDenPyzzKATDerxyq7g81fjWRE6Nknu346uFFvSvaCYiYsLY",
            "AUTH_COOKIE_KEY":"eyJsb2dpbklkIjoieWFvemhlbmd3ZWkiLCJuYW1lIjoi5aea5q2j5aiBIiwicmVxdWVzdElwIjoiMTAuMTI5LjIwLjE3MSIsInNpZ24iOiJlZTVhMGY1MmM1ZDgwNDA0NDc2NmIwOGZiMzE5MDU0ODUzOWY2Nzg3YTJlZTU3NDhlNTQ3Mzc2MTIxOGNlZGQyIiwidGltZSI6MTYxNjM5ODU.",
            "AUTH_COOKIE_KEY58":"dK99Wt571v6xkSzU5vYDJZiU6K2bJrF4QNdhjsbnxhVJqajKbMS31StA8H8QL3PCTRYeTvCTpk8X1DLqFKptMNSLpNjgdLHmCaMG9uYX1ffmxSrAKKCv3N272sMJh2C1fU3hArJ2CT2sVJRJfjkTFjhDd6HQmdirxcW1C5L4g4LBP8x3tFD3DqdBNyFE9mkJcG2HWRuGLx8GWcC31mmVcyijGZcv4CVnHgz8",
            "JSESSIONID":"aaaUn65z1Ah2veRLp3yHx",
            "SESSIONID":"838bd542-a167-4086-a780-d6a88c6d2d51"}
    
    if user=='孙世江':
        cookies=cookies1
    elif user=='马宇星':
        cookies=cookies2
    elif user=='罗海洋':
        cookies=cookies3
    elif user=='叶星星':
        cookies=cookies4
    elif user=='王欣欣':
        cookies=cookies5
    elif user=='龙帅':
        cookies=cookies6
    elif user=='白云涧':
        cookies=cookies7
    elif user=='姚正威':
        cookies=cookies8


    #接口请求发送获取当前月份响应数据
    url='https://oa.sogou-inc.com/attd/api/attd/mine/attendance/events'

    data={
    'month':month_ip
    }
    h=requests.post(url=url,headers=headers,data=data,cookies=cookies)

    #获取当前登录用户信息的name信息，以输出姓名
    url_login='https://oa.sogou-inc.com/attd/api/login-info'

    data={
    'month':month_ip
    }
    h_login=requests.post(url=url_login,headers=headers,data=data,cookies=cookies)
    user_ip=h_login.json().get('data').get('name') 

    

    #根据返回的响应数据筛选需要爬取的内容
    kqlist=h.json().get('data')
    if len(kqlist)!=0:
        toast_if='导出成功'
        l1=[]                                                                    #生成当前月份所有的考勤记录信息
        for i in kqlist:
            l2=[]                                                                #生成当前月份每一天的考勤记录信息
            
            current_date=i.get('start')                                          #获取当前日期（仅年月日）
            time1=datetime.datetime.strptime(current_date, '%Y-%m-%d')   
            time2=current_date.split('-')[1]                                     #获取当前月份
            totalhour=i.get('data').get('totalHour')                             #获取当前总的打卡时间,如：9.98h
            totalhour_hour=int(totalhour)
            totalhour_mins=int((totalhour%1)*60)
            total_time=f'{totalhour_hour}小时{totalhour_mins}分'                 #将获取到的当前总的打卡时间转化格式，如：9小时58分
            #date_date=current_date + get_week_day(time1)                         #获取当前日期（年月日 + 星期）
            date_date=current_date.replace('-','/') + get_week_day(time1) 
            if 'startTime' in i.get('data').keys():                              #判断如果当前日期存在打卡时间
                start=(i.get('data').get('startTime'))/1000                      #带微秒格式的时间戳/1000=不带微秒格式的时间戳（首次打卡时间）
                end=(i.get('data').get('endTime'))/1000                          #带微秒格式的时间戳/1000=不带微秒格式的时间戳（末次打卡时间）
                startime=time.strftime('%Y/%m/%d %H:%M',time.localtime(start))   #带微秒格式的时间戳转化为日期格式（首次打卡时间）
                endtime=time.strftime('%Y/%m/%d %H:%M',time.localtime(end))      #带微秒格式的时间戳转化为日期格式（末次打卡时间）
                if totalhour >=11:                                               #判断当前日期总打卡时间是否>11小时
                    l2=[time2+'月',user_ip,date_date,startime,endtime,total_time,20,'','加班补助']
                else:
                    l2=[time2+'月',user_ip,date_date,startime,endtime,total_time,'','','']
            else:               
                startime=''
                endtime=''
                total_time=''
                l2=[time2+'月',user_ip,date_date,startime,endtime,total_time,'','','']

            l1.append(l2)
        print(l1)




        #本次创建excel表并将爬取的数据填写到表格中
        wo = Workbook()
        wsheet1 = wo.active
        wsheet1['A1']='月份'            #生成表头
        wsheet1['B1']='姓名'
        wsheet1['C1']='日期'
        wsheet1['D1']='首次打卡时间'
        wsheet1['E1']='末次打卡时间'
        wsheet1['F1']='考勤总时长'
        wsheet1['G1']='餐费补助'
        wsheet1['H1']='打车费'
        wsheet1['I1']='备注'


        for i in range(0,len(l1)):
            for j in range(0,len(l1[i])):
                 wsheet1.cell(column=j+1,row=i+2).value=l1[i][j]
        wsheet1.merge_cells(f'A2:A{len(l1)+1}')       #将A列（月份）除表头外的单元格进行合并
        wsheet1.cell(2,2).value=time2+'月'            #将A列（月份）列表合并单元格后显示：月份
        wsheet1.merge_cells(f'B2:B{len(l1)+1}')       #将B列（姓名）除表头外的单元格进行合并
        wsheet1.cell(2,2).value=user_ip              #将B列（姓名）列表合并单元格后显示：姓名

        align=Alignment(horizontal='center',vertical='center')   #定义表格格式：对齐（居中）
        font=Font('宋体',bold=True)                              #定义字体格式：宋体+加粗



        wsheet1['A2'].alignment=align                            #将A列（月份）数据内容居中显示
        wsheet1['B2'].alignment=align                            #将B列（月份）数据内容居中显示



        for i in ['A','B','C','D','E','F','G','H','I']:
            wsheet1[f'{i}1'].font=font                           #将第一行（表头）数据内容宋体+加粗显示
        wsheet1.column_dimensions['C'].width=20                  #定义单元格格式：自定义宽度
        wsheet1.column_dimensions['D'].width=20
        wsheet1.column_dimensions['E'].width=20
        wsheet1.column_dimensions['F'].width=12
        for i in range(1,len(l1)+3):                             
            wsheet1[f"G{i}"].alignment=align                     #将G列（餐补）所有数据内容居中显示
            wsheet1[f"I{i}"].alignment=align                     #将I列（备注）所有数据内容居中显示

        wsheet1[f"A{len(l1)+2}"]='总计'                          #最后一行第一列显示‘总计’
        
        for row in wsheet1.rows:
            for cell in row:
                cell.border = Border(top = Side(border_style='thin',color='FF000000'), right = Side(border_style='thin', color='FF000000'), bottom = Side(border_style='thin', color='FF000000'),left = Side(border_style='thin', color='FF000000'))
        wo.save('考勤.xlsx')

        import pandas as pd
        df=pd.read_excel('考勤.xlsx')
        sum_col=int(df['餐费补助'].sum())                        #计算G列（餐补）所有数据内容求和
        wsheet1[f"G{len(l1)+2}"]=sum_col
        wo.save('考勤.xlsx')
    else:
        toast_if='导出失败'
        
    
    '''--------------------------------------------------------------------------------------------------'''
    action.configure(text=toast_if)  
     

action = ttk.Button(mighty, text="Click Me!", command=click_me)   
action.grid(column=2, row=4)                                

ttk.Label(mighty, text="Choose a user:").grid(column=1, row=1)
number1 = tk.StringVar()
number1_chosen = ttk.Combobox(mighty, width=12, textvariable=number1, state='readonly')
number1_chosen['values'] = ('马宇星','孙世江','罗海洋','叶星星','王欣欣','龙帅','白云涧','姚正威')
number1_chosen.grid(column=1, row=2)
number1_chosen.current(0)

ttk.Label(mighty, text="Choose a year:").grid(column=1, row=3)
number2 = tk.StringVar()
number2_chosen = ttk.Combobox(mighty, width=12, textvariable=number2, state='readonly')
number2_chosen['values'] = ('2021', '2022', '2023','2024', '2025', '2026','2027', '2028', '2029')
number2_chosen.grid(column=1, row=4)
number2_chosen.current(0)

ttk.Label(mighty, text="Choose a month:").grid(column=1, row=5)
number3 = tk.StringVar()
number3_chosen = ttk.Combobox(mighty, width=12, textvariable=number3, state='readonly')
number3_chosen['values'] = ('01','02', '03', '04', '05','06','07','08','09','10','11','12')
number3_chosen.grid(column=1, row=6)
number3_chosen.current(0)

def _quit():
    win.quit()
    win.destroy()
    exit() 

win.mainloop()

