import requests
import selenium
from selenium import webdriver
import time
import datetime
import json
from interval import Interval
from urllib.parse import urlencode
from openpyxl import Workbook
from openpyxl.styles import Font,Color,Alignment
headers = {
    'Host' : 'oa.sogou-inc.com',
    'Referer' : 'https://oa.sogou-inc.com/statics/attd/web2/index.html',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.90 Safari/537.36 Edg/89.0.774.54'
}

#根据日期返回星期几
def get_week_day(date):
  week_day_dict = {
    0 : '星期一',
    1 : '星期二',
    2 : '星期三',
    3 : '星期四',
    4 : '星期五',
    5 : '星期六',
    6 : '星期天',
  }
  day = date.weekday()
  return week_day_dict[day]


#selenium自动登录
post={}
url1="https://oa.sogou-inc.com/"
driver = webdriver.Edge(executable_path=r'E:\Python\Scripts\msedgedriver.exe')
driver.get(url1)
time.sleep(3)
driver.find_element_by_id("xiaopFace").click()
time.sleep(3)
driver.execute_script("window.open('https://oa.sogou-inc.com/statics/attd/web2/index.html');")
time.sleep(3)

#登陆后获取cookie
cookie_items=driver.get_cookies()
for cookie_item in cookie_items:
    post[cookie_item['name']] = cookie_item['value']
cookie_str = json.dumps(post)
with open('cookie.txt', 'w+') as f:
    f.write(cookie_str)
print("cookies信息已保存到本地")

with open('cookie.txt', 'r') as f:
    cookie = f.read()
cookies = json.loads(cookie)


#接口请求发送获取当前月份响应数据
url='https://oa.sogou-inc.com/attd/api/attd/mine/attendance/events'

data={
'month':'2021-02'
}
h=requests.post(url=url,headers=headers,data=data,cookies=cookies)



#根据返回的响应数据筛选需要爬取的内容
kqlist=h.json().get('data')              
l1=[]                                                                    #生成当前月份所有的考勤记录信息
for i in kqlist:
    l2=[]                                                                #生成当前月份每一天的考勤记录信息
    
    current_date=i.get('start')                                          #获取当前日期（仅年月日）
    time1=datetime.datetime.strptime(current_date, '%Y-%m-%d')   
    time2=current_date.split('-')[1]                                     #获取当前月份
    totalhour=i.get('data').get('totalHour')                             #获取当前总的打卡时间,如：9.98h
    totalhour_hour=int(totalhour)
    totalhour_mins=int((totalhour%1)*60)
    total_time=f'{totalhour_hour}小时{totalhour_mins}分'                 #将获取到的当前总的打卡时间转化格式，如：9小时58分
    date_date=current_date + get_week_day(time1)                         #获取当前日期（年月日 + 星期）
    if 'startTime' in i.get('data').keys():                              #判断如果当前日期存在打卡时间
        start=(i.get('data').get('startTime'))/1000                      #带微秒格式的时间戳/1000=不带微秒格式的时间戳（首次打卡时间）
        end=(i.get('data').get('endTime'))/1000                          #带微秒格式的时间戳/1000=不带微秒格式的时间戳（末次打卡时间）
        startime=time.strftime('%Y/%m/%d %H:%M',time.localtime(start))   #带微秒格式的时间戳转化为日期格式（首次打卡时间）
        endtime=time.strftime('%Y/%m/%d %H:%M',time.localtime(end))      #带微秒格式的时间戳转化为日期格式（末次打卡时间）
        if totalhour >=11:                                               #判断当前日期总打卡时间是否>11小时
            l2=[time2+'月','孙世江',date_date,startime,endtime,total_time,20,'','加班补助']
        else:
            l2=[time2+'月','孙世江',date_date,startime,endtime,total_time,'','','']
    else:               
        startime=''
        endtime=''
        total_time=''
        l2=[time2+'月','孙世江',date_date,startime,endtime,total_time,'','','']

    l1.append(l2)
print(l1)



#本次创建excel表并将爬取的数据填写到表格中
wo = Workbook()
wsheet1 = wo.active
wsheet1['A1']='月份'            #生成表头
wsheet1['B1']='姓名'
wsheet1['C1']='日期'
wsheet1['D1']='首次打卡时间'
wsheet1['E1']='末次打卡时间'
wsheet1['F1']='考勤总时长'
wsheet1['G1']='餐费补助'
wsheet1['H1']='打车费'
wsheet1['I1']='备注'


for i in range(0,len(l1)):
    for j in range(0,len(l1[i])):
         wsheet1.cell(column=j+1,row=i+2).value=l1[i][j]
wsheet1.merge_cells(f'A2:A{len(l1)+1}')       #将A列（月份）除表头外的单元格进行合并
wsheet1.cell(2,2).value=time2+'月'            #将A列（月份）列表合并单元格后显示：月份
wsheet1.merge_cells(f'B2:B{len(l1)+1}')       #将B列（姓名）除表头外的单元格进行合并
wsheet1.cell(2,2).value='孙世江'              #将B列（姓名）列表合并单元格后显示：姓名

align=Alignment(horizontal='center',vertical='center')   #定义表格格式：对齐（居中）
font=Font('宋体',bold=True)                              #定义字体格式：宋体+加粗



wsheet1['A2'].alignment=align                            #将A列（月份）数据内容居中显示
wsheet1['B2'].alignment=align                            #将B列（月份）数据内容居中显示



for i in ['A','B','C','D','E','F','G','H','I']:
    wsheet1[f'{i}1'].font=font                           #将第一行（表头）数据内容宋体+加粗显示
wsheet1.column_dimensions['C'].width=20                  #定义单元格格式：自定义宽度
wsheet1.column_dimensions['D'].width=20
wsheet1.column_dimensions['E'].width=20
wsheet1.column_dimensions['F'].width=12
for i in range(1,len(l1)+3):                             
    wsheet1[f"G{i}"].alignment=align                     #将G列（餐补）所有数据内容居中显示
    wsheet1[f"I{i}"].alignment=align                     #将I列（备注）所有数据内容居中显示

wsheet1[f"A{len(l1)+2}"]='总计'                          #最后一行第一列显示‘总计’

wo.save('考勤.xlsx')

import pandas as pd
df=pd.read_excel('考勤.xlsx')
sum_col=int(df['餐费补助'].sum())                        #计算G列（餐补）所有数据内容求和
wsheet1[f"G{len(l1)+2}"]=sum_col
wo.save('考勤.xlsx')